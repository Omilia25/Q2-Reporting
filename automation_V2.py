import openpyxl

source_wb = openpyxl.load_workbook('source.xlsx')
vouchers_wb = openpyxl.load_workbook('template.xlsx')

data_sheet = source_wb['Sheet1']
template_sheet = vouchers_wb['Sheet1']

sum_rows = data_sheet.max_row
for i in range(1, sum_rows + 1):
        
        
        template_sheet['B29'].value = data_sheet['A'+str(i)].value 
        template_sheet['H29'].value = data_sheet['F'+str(i)].value 
        template_sheet['E21'].value = data_sheet['E'+str(i)].value 
        template_sheet['H19'].value = data_sheet['B'+str(i)].value 
        template_sheet['H18'].value = data_sheet['C'+str(i)].value 
        template_sheet['H14'].value = data_sheet['B'+str(i)].value 
        template_sheet['H12'].value = data_sheet['B'+str(i)].value 
        template_sheet['H6'].value = data_sheet['F'+str(i)].value 
        template_sheet['B10'].value = data_sheet['D'+str(i)].value 
        template_sheet['D5'].value = data_sheet['A'+str(i)].value 
        template_sheet['E3'].value = data_sheet['E'+str(i)].value 
        template_sheet['H2'].value = data_sheet['B'+str(i)].value 
        template_sheet['H1'].value = data_sheet['C'+str(i)].value 

        template_sheet.column_dimensions['E'].width = 20
        template_sheet.column_dimensions['H'].width = 20
        template_sheet['h2'].number_format = "DD/MM/YYYY"
        template_sheet['h12'].number_format = "DD/MM/YYYY"
        template_sheet['h14'].number_format = "DD/MM/YYYY"
        template_sheet['h19'].number_format = "DD/MM/YYYY"

        # saving the generated sheet
        vouchers_wb.save('./Q2_Vouchers/Q2_voucher_' +str(i)+ '.xlsx')

        # deleting the template worksheet before saving
        del_template = openpyxl.load_workbook('./Q2_Vouchers/Q2_voucher_' +str(i)+ '.xlsx')
        del del_template['template'] 

        del_template.save('./Q2_Vouchers/Q2_voucher_' +str(i)+ '.xlsx')

print('Succes!')