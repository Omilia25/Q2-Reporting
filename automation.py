from re import template
import openpyxl

source_wb = openpyxl.load_workbook('source.xlsx')
vouchers_wb = openpyxl.load_workbook('template.xlsx')

data_sheet = source_wb['Sheet1']
template_sheet = vouchers_wb['Sheet1']

sum_rows = data_sheet.max_row
for i in range(1, sum_rows + 1):

        gen_sheet = vouchers_wb.create_sheet()

        # Calculate total rows and columns in source
        total_rows = template_sheet.max_row
        total_columns = template_sheet.max_column

        # copying the template
        for x in range(1, total_rows + 1):
            for j in range(1, total_columns + 1):
                # reading cell values from excel file
                active_cell = template_sheet.cell(row = x, column = j)

                # Writing the read value to destination excel file
                gen_sheet.cell(row=x, column=j).value = active_cell.value

        gen_sheet['B29'].value = data_sheet['A' +str(i)].value 
        gen_sheet['H29'].value = data_sheet['F' +str(i)].value 
        gen_sheet['E21'].value = data_sheet['E'+str(i)].value 
        gen_sheet['H19'].value = data_sheet['B'+str(i)].value 
        gen_sheet['H18'].value = data_sheet['C'+str(i)].value 
        gen_sheet['H14'].value = data_sheet['B'+str(i)].value 
        gen_sheet['H12'].value = data_sheet['B'+str(i)].value 
        gen_sheet['H6'].value = data_sheet['F'+str(i)].value 
        gen_sheet['B10'].value = data_sheet['D'+str(i)].value 
        gen_sheet['D5'].value = data_sheet['A'+str(i)].value 
        gen_sheet['E3'].value = data_sheet['E'+str(i)].value 
        gen_sheet['H2'].value = data_sheet['B'+str(i)].value 
        gen_sheet['H1'].value = data_sheet['C'+str(i)].value 

        gen_sheet.column_dimensions['E'].width = 15
        gen_sheet.column_dimensions['H'].width = 15
        gen_sheet['h2'].number_format = "DD/MM/YYYY"
        gen_sheet['h12'].number_format = "DD/MM/YYYY"
        gen_sheet['h14'].number_format = "DD/MM/YYYY"
        gen_sheet['h19'].number_format = "DD/MM/YYYY"

        # saving the generated sheet
        vouchers_wb.save('template.xlsx')

print(vouchers_wb.sheetnames)